import os
import io
import math
import json
import re
import logging
import asyncio
import pandas as pd
import openpyxl
from openpyxl.cell.cell import MergedCell
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime

from dotenv import load_dotenv
from google import genai
from google.genai import types

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("extraction_service")


# ============================================================================
# CONFIGURATION
# ============================================================================
MODEL_NAME = "gemini-2.0-flash"
MAX_HEADER_ROWS = 40   # Rows sent to LLM for canonical field extraction
LINE_ITEM_KEYWORDS = ["part", "qty", "quantity", "price", "asn", "po", "amount", "value", "description"]

# Centralized configuration for Header Fields
# Add synonyms here to automatically update the LLM prompt
CANONICAL_FIELDS_CONFIG = {
    "supplier_name": {
        "description": "Name of the vendor, shipper, exporter, or seller.",
        "synonyms": ["Shipper", "Vendor", "Supplier", "Exporter", "Seller"]
    },
    "supplier_location": {
        "description": "Full physical address of the supplier.",
        "synonyms": ["Address", "Origin Address", "Address of Shipper"]
    },
    "invoice_number": {
        "description": "The unique invoice identifier string.",
        "synonyms": ["Inv#", "Invoice No.", "Invoice Number", "No.", "Reference"]
    },
    "invoice_date": {
        "description": "Date of invoice issuance (normalized to YYYY-MM-DD).",
        "synonyms": ["Date", "Invoice Date", "Issue Date", "Date of Issue"]
    },
    "currency": {
        "description": "The ISO 3-letter currency code (e.g., USD, EUR, INR).",
        "synonyms": ["Currency", "Currency Code", "Monetary Unit"]
    },
    "Incoterm": {
        "description": "Delivery and payment terms (e.g., FCA, EXW, FOB).",
        "synonyms": ["Incoterm", "Delivery Terms", "Terms of Delivery"]
    },
    "commercial_invoice_value": {
        "description": "Total monetary value of the commercial invoice.",
        "synonyms": ["Total Value", "Invoice Amount", "Calculated Total", "Grand Total","SUB TOTAL","GRAN TOTAL"]
    },
    "HAWB_number": {
        "description": "House Air Waybill number.",
        "synonyms": ["HAWB", "HAWB#", "House AWB", "Air Waybill","HAWB#:"]   
    },
    "MAWB_number": {
        "description": "Master Air Waybill number.",
        "synonyms": ["MAWB", "MAWB#", "Master AWB"]
    }
}

# Centralized configuration for Table Column mapping
# Add synonyms here to improve Python's column detection
LINE_ITEMS_MAPPING = {
    "ASN":              ["asn", "asn#", "asn no", "serial number", "pack list"],
    "part_number":      ["part", "part number", "part no", "part no.", "pin", "sku number", "sku", "item", "material", "model"],
    "PO":               ["po", "po#", "po number", "purchase order", "order no", "s/o"],
    "quantity":         ["qty", "q'ty", "quantity", "units", "pcs"],
    "unit_price":       ["unit price", "unit cost", "price", "rate", "unit value"],
    "total_value":      ["total", "amount", "value", "ext price", "extended"],
    "country_of_origin":["country", "origin", "coo", "made in", "city of origin"],
}

_PROJECT_ID = os.getenv("GCP_PROJECT_ID")
_LOCATION = os.getenv("GCP_LOCATION", "us-central1")

if not _PROJECT_ID:
    raise EnvironmentError("GCP_PROJECT_ID environment variable is required but not set.")

_client = genai.Client(vertexai=True, project=_PROJECT_ID, location=_LOCATION)

# Currency normalization map
CURRENCY_MAP = {
    "us dollars": "USD", "usd": "USD", "$": "USD",
    "euro": "EUR", "euros": "EUR", "eur": "EUR", "€": "EUR",
    "gbp": "GBP", "pounds": "GBP", "£": "GBP",
    "inr": "INR", "rupees": "INR", "₹": "INR",
    "cad": "CAD", "aud": "AUD", "jpy": "JPY", "cny": "CNY", "sgd": "SGD",
    "mxn": "MXN", "brl": "BRL",
}


def retry_with_backoff(func, max_retries=3, initial_delay=2):
    """Decorator for exponential backoff retries on Gemini API calls."""
    async def wrapper(*args, **kwargs):
        delay = initial_delay
        last_exception = None
        for attempt in range(max_retries + 1):
            try:
                # If the function is sync, we run it in a thread
                if not asyncio.iscoroutinefunction(func):
                    return await asyncio.to_thread(func, *args, **kwargs)
                return await func(*args, **kwargs)
            except Exception as e:
                last_exception = e
                # Retry on quota (429) or service (5xx) errors
                if attempt < max_retries and ("429" in str(e) or "503" in str(e) or "500" in str(e)):
                    logger.warning(f"API Attempt {attempt+1} failed ({e}). Retrying in {delay}s...")
                    await asyncio.sleep(delay)
                    delay *= 2
                else:
                    break
        raise last_exception
    return wrapper


def build_system_instruction() -> str:
    """Build system instruction dynamically from CANONICAL_FIELDS_CONFIG."""
    fields_info = []
    for field, config in CANONICAL_FIELDS_CONFIG.items():
        synonyms_str = ", ".join(config["synonyms"])
        fields_info.append(f"- {field}: {config['description']} (Synonyms: {synonyms_str})")
    
    fields_list = "\n    ".join(fields_info)

    return f"""
    You are a strict Invoice Data Extractor for an Enterprise ERP system.

    The data is a list of Excel cells in format: CELL_REF: VALUE
    Example: I1: SANMINA CORPORATION
             D7: INV-2024-8821

    You MUST use the exact CELL_REF shown in the data for the cell_ref of each extracted field.

    STEP 1: STRICT DOCUMENT CLASSIFICATION
    Set is_invoice=true ONLY if BOTH:
    - Header area (first 15 rows) explicitly contains "INVOICE" or "COMMERCIAL INVOICE"
    - A valid non-empty INVOICE NUMBER can be extracted
    Otherwise is_invoice=false. Return empty canonical_fields if false.

    STEP 2: EXTRACT CANONICAL FIELDS
    For each field, return:
    {{
      "value": <extracted value or null>,
      "cell_ref": <exact cell ref from data e.g. "I1">,
      "confidence": <0.0-1.0>
    }}
    Confidence: 1.0=exact label match, 0.7-0.9=heuristic, 0.5-0.7=ambiguous, 0.0=not found

    Fields Definitions:
    {fields_list}

    STEP 3: LINE ITEMS TABLE CONFIDENCE
    Return a single 0.0-1.0 score for how reliably you believe the line items table exists in this sheet.

    STRICT OUTPUT: JSON only, no markdown, no code blocks.
    """


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def normalize_date(value: Any) -> Optional[str]:
    """Normalize date to YYYY-MM-DD, stripping timestamps."""
    if not value:
        return None
    s = str(value).strip()
    # Already clean date
    match = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    if match:
        return match.group(1)
    # Try pandas parse
    try:
        return pd.to_datetime(s).strftime("%Y-%m-%d")
    except Exception:
        return s


def normalize_currency(value: Any) -> Optional[str]:
    """Normalize currency to ISO 3-letter code."""
    if not value:
        return None
    key = str(value).strip().lower()
    return CURRENCY_MAP.get(key, str(value).strip().upper())


def compute_overall_confidence(canonical_fields: Dict[str, Any]) -> float:
    """
    Compute overall confidence as average of all _confidence fields
    where a value was found (skips null/missing fields).
    """
    scores = []
    for key, val in canonical_fields.items():
        if key.endswith("_confidence") and val is not None:
            field_name = key.replace("_confidence", "")
            if canonical_fields.get(field_name) is not None:
                scores.append(float(val))
    if not scores:
        return 0.0
    return round(sum(scores) / len(scores), 3)


def remove_null_values(data: Any) -> Any:
    """Recursively remove None/empty/NaN values."""
    if isinstance(data, dict):
        return {
            k: remove_null_values(v)
            for k, v in data.items()
            if v not in [None, "", "None", "null", []]
            and not (isinstance(v, float) and math.isnan(v))
        }
    if isinstance(data, list):
        return [
            remove_null_values(i)
            for i in data
            if i not in [None, "", "None", "null"]
            and not (isinstance(i, float) and math.isnan(i))
        ]
    return data


# ============================================================================
# 1. openpyxl — CELL MAP BUILDER
# ============================================================================
def build_cell_map(ws) -> List[Tuple[str, str, Any]]:
    """
    Build a list of (cell_ref, col_letter, value) tuples from an openpyxl worksheet.
    Rules:
    - Skip hidden rows and hidden columns
    - Skip MergedCell placeholders (keep only the parent/top-left cell of each merge)
    - Returns: list of (ref, col_letter, value) sorted by row then column
    """
    # Collect hidden column letters
    hidden_cols = set()
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.hidden:
            hidden_cols.add(col_letter)

    cell_map = []
    for row in ws.iter_rows():
        row_dim = ws.row_dimensions.get(row[0].row)
        if row_dim and row_dim.hidden:
            continue  # Skip hidden rows

        for cell in row:
            if isinstance(cell, MergedCell):
                continue  # Skip merge placeholders — only parent cell has value
            col_letter = cell.column_letter
            if col_letter in hidden_cols:
                continue  # Skip hidden columns
            if cell.value is None or str(cell.value).strip() == "":
                continue  # Skip empty cells
            cell_map.append((cell.coordinate, col_letter, cell.value))

    return cell_map


def cell_map_to_text(cell_map: List[Tuple], max_rows: int = MAX_HEADER_ROWS) -> str:
    """
    Convert cell map to LLM-readable text with explicit cell references.
    Format: "I1: SANMINA CORPORATION\nD7: INV-2024-8821\n..."
    Limited to first `max_rows` rows of data.
    """
    if not cell_map:
        return ""
    # Get unique row numbers from cell_map
    rows_seen = []
    for ref, col, val in cell_map:
        row_num = int("".join(filter(str.isdigit, ref)))
        if row_num not in rows_seen:
            rows_seen.append(row_num)
        if len(rows_seen) >= max_rows:
            break

    cutoff_row = rows_seen[-1] if rows_seen else max_rows
    lines = []
    for ref, col, val in cell_map:
        row_num = int("".join(filter(str.isdigit, ref)))
        if row_num > cutoff_row:
            break
        lines.append(f"{ref}: {val}")
    return "\n".join(lines)


# ============================================================================
# 2. LINE ITEMS TABLE DETECTION + pandas reading
# ============================================================================
def detect_table_start_row(cell_map: List[Tuple]) -> Optional[int]:
    """
    Detect which row is the line items table header by scanning for keyword density.
    Returns the 1-indexed row number of the header row.
    """
    row_scores: Dict[int, int] = {}
    for ref, col, val in cell_map:
        row_num = int("".join(filter(str.isdigit, ref)))
        val_lower = str(val).lower()
        score = sum(1 for kw in LINE_ITEM_KEYWORDS if kw in val_lower)
        if score > 0:
            row_scores[row_num] = row_scores.get(row_num, 0) + score

    if not row_scores:
        return None
    return max(row_scores, key=row_scores.get)


def build_line_items_with_refs(file_bytes: bytes, sheet_name: str, header_row: int) -> List[Dict[str, Any]]:
    """
    Use pandas to read the line items table starting from header_row.
    Compute exact cell_ref (col_letter + row_number) for every field in every row.
    """
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row - 1,   # pandas is 0-indexed
        dtype=str
    )
    df = df.dropna(how='all').reset_index(drop=True)
    df = df.fillna("")

    # Map canonical field names to likely column name patterns
    FIELD_PATTERNS = LINE_ITEMS_MAPPING

    # Detect which DataFrame column maps to which canonical field
    col_to_field: Dict[str, str] = {}
    field_to_col: Dict[str, str] = {}  # canonical field → df column name
    for col in df.columns:
        col_lower = str(col).lower().strip()
        for field, patterns in FIELD_PATTERNS.items():
            if field not in field_to_col and any(p in col_lower for p in patterns):
                col_to_field[col] = field
                field_to_col[field] = col
                break

    # Build openpyxl column letter map: df column index → Excel column letter
    # pandas header_row in Excel = header_row, data starts at header_row + 1
    wb_temp = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws_temp = wb_temp[sheet_name]

    # Find the actual Excel column letters by matching header values
    header_excel_row = ws_temp[header_row]
    col_letter_map: Dict[str, str] = {}  # df column name → Excel column letter
    for cell in header_excel_row:
        if cell.value and str(cell.value).strip():
            col_name = str(cell.value).strip()
            for df_col in df.columns:
                if str(df_col).strip() == col_name:
                    col_letter_map[df_col] = cell.column_letter
                    break

    # Numeric fields that MUST have data for a row to be a real line item
    NUMERIC_FIELDS = {"quantity", "unit_price", "total_value"}
    # How many consecutive non-data rows before we stop reading
    MAX_CONSECUTIVE_EMPTY = 2

    line_items = []
    consecutive_non_data = 0

    for idx, row in df.iterrows():
        excel_row_num = header_row + 1 + int(idx)

        # Check if this row has ANY numeric field value
        has_numeric_data = False
        for field in NUMERIC_FIELDS:
            df_col = field_to_col.get(field)
            if df_col:
                raw = str(row.get(df_col, "")).strip()
                if raw and raw.lower() not in ("none", "nan", ""):
                    try:
                        float(raw.replace(",", "").replace("$", "").strip())
                        has_numeric_data = True
                        break
                    except ValueError:
                        pass

        if not has_numeric_data:
            consecutive_non_data += 1
            # Stop reading once we've seen enough consecutive non-data rows
            if consecutive_non_data >= MAX_CONSECUTIVE_EMPTY:
                logger.info(f"Table end detected at Excel row {excel_row_num} "
                            f"({MAX_CONSECUTIVE_EMPTY} consecutive non-data rows). Stopping.")
                break
            continue  # Skip this row but keep scanning for now

        # Valid data row — reset counter
        consecutive_non_data = 0

        item: Dict[str, Any] = {"row_ref": excel_row_num}

        for field, df_col in field_to_col.items():
            raw_val = row.get(df_col, "")
            val = str(raw_val).strip() if raw_val != "" else None
            col_letter = col_letter_map.get(df_col, "")
            cell_ref = f"{col_letter}{excel_row_num}" if col_letter else None

            # Type coerce numeric fields
            if field in NUMERIC_FIELDS and val:
                try:
                    val = float(str(val).replace(",", "").replace("$", "").strip())
                except ValueError:
                    pass

            item[field] = val
            if cell_ref:
                item[f"{field}_ref"] = cell_ref

        line_items.append(item)

    return line_items


# ============================================================================
# 3. SHEET READER (Entry Point for Both Paths)
# ============================================================================
def extract_sheet_data(file_bytes: bytes, filename: str) -> Dict[str, Dict[str, Any]]:
    """
    Read all non-empty sheets using openpyxl for cell map and metadata.
    Returns dict of sheet_name → {cell_map, cell_text, total_rows, file_bytes, sheet_name}
    """
    logger.info(f"Reading sheets from: {filename}")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        if "password" in str(e).lower() or "encrypted" in str(e).lower():
            raise ValueError(f"File '{filename}' is password protected or encrypted.")
        raise ValueError(f"Failed to read Excel file '{filename}': {str(e)}")

    sheets = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cell_map = build_cell_map(ws)
        if not cell_map:
            continue
        sheets[sheet_name] = {
            "sheet_name": sheet_name,
            "cell_map": cell_map,
            "cell_text": cell_map_to_text(cell_map, MAX_HEADER_ROWS),
            "total_rows": ws.max_row,
            "file_bytes": file_bytes,
        }
    return sheets


# ============================================================================
# 4. LLM CALL (BLOCKING)
# ============================================================================
def _call_gemini_blocking(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """
    Blocking Gemini call for canonical field extraction.
    Always wrap with asyncio.to_thread in async contexts.
    """
    cell_text = metadata["cell_text"]
    sheet_name = metadata["sheet_name"]

    if not cell_text.strip():
        raise ValueError(f"No cell data for sheet '{sheet_name}'")

    system_instruction = build_system_instruction()

    prompt = f"""Analyze sheet: "{sheet_name}"

CELL DATA:
{cell_text}

Extract invoice fields with exact cell references and confidence scores.
"""

    field_schema = {
        "type": "object",
        "properties": {
            "value": {"type": "string"},
            "cell_ref": {"type": "string"},
            "confidence": {"type": "number"}
        }
    }
    numeric_field_schema = {
        "type": "object",
        "properties": {
            "value": {"type": "number"},
            "cell_ref": {"type": "string"},
            "confidence": {"type": "number"}
        }
    }

    # Build properties for canonical_fields dynamically
    canonical_properties = {}
    for field in CANONICAL_FIELDS_CONFIG:
        # Check if the field likely contains numeric data
        is_numeric = any(kw in field.lower() for kw in ["value", "amount", "total", "price"])
        canonical_properties[field] = numeric_field_schema if is_numeric else field_schema

    output_schema = {
        "type": "object",
        "properties": {
            "is_invoice": {"type": "boolean"},
            "document_type": {"type": "string"},
            "canonical_fields": {
                "type": "object",
                "properties": canonical_properties
            },
            "line_items_confidence": {"type": "number"}
        },
        "required": ["is_invoice", "document_type", "canonical_fields", "line_items_confidence"]
    }

    logger.info(f"Calling Gemini ({MODEL_NAME}) for sheet '{sheet_name}'")
    response = _client.models.generate_content(
        model=MODEL_NAME,
        contents=prompt,
        config=types.GenerateContentConfig(
            system_instruction=system_instruction,
            response_mime_type="application/json",
            response_schema=output_schema,
            temperature=0.1,
        )
    )

    return response.parsed if response.parsed else json.loads(response.text)


# ============================================================================
# 5. POST-PROCESS: Flatten to BigQuery-friendly format
# ============================================================================
def flatten_canonical_fields(raw_canonical: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert nested {value, cell_ref, confidence} objects to flat fields:
    supplier_name, supplier_name_ref, supplier_name_confidence
    Also normalize date and currency values.
    """
    flat = {}
    for field_name, field_data in raw_canonical.items():
        if not isinstance(field_data, dict):
            continue
        val = field_data.get("value")
        ref = field_data.get("cell_ref")
        conf = field_data.get("confidence")

        # Normalize specific fields
        if field_name == "invoice_date" and val:
            val = normalize_date(val)
        if field_name == "currency" and val:
            val = normalize_currency(val)

        if val is not None:
            flat[field_name] = val
        if ref:
            flat[f"{field_name}_ref"] = ref
        if conf is not None:
            flat[f"{field_name}_confidence"] = round(float(conf), 3)

    return flat


# ============================================================================
# 6. ASYNC WRAPPER WITH RETRY
# ============================================================================
@retry_with_backoff
async def process_sheet_async(metadata: Dict[str, Any]) -> Dict[str, Any]:
    """Run blocking Gemini call in a thread with retry logic."""
    try:
        return await asyncio.to_thread(_call_gemini_blocking, metadata)
    except Exception as e:
        logger.error(f"Final failure for sheet '{metadata.get('sheet_name')}': {e}")
        return {
            "is_invoice": False,
            "document_type": "ERROR",
            "error": str(e),
            "canonical_fields": {},
            "line_items_confidence": 0.0
        }


# ============================================================================
# 7. SYNC ENTRY POINT
# ============================================================================
def extract_invoice_from_bytes(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    logger.info(f"Sync extraction for: {filename}")

    sheets = extract_sheet_data(file_bytes, filename)
    if not sheets:
        return {"_file_info": {"filename": filename, "error": "No data found"}}

    results = {}
    skipped = []
    processed = 0

    for name, meta in sheets.items():
        top_text = meta["cell_text"].lower()
        if not any(kw in top_text for kw in ["invoice", "inv.", "commercial invoice"]):
            skipped.append({"name": name, "reason": "No invoice keyword"})
            continue

        try:
            llm_result = _call_gemini_blocking(meta)
            if not llm_result.get("is_invoice"):
                skipped.append({"name": name, "reason": f"LLM: {llm_result.get('document_type', 'Not Invoice')}"})
                continue

            # Flatten canonical fields
            canonical = flatten_canonical_fields(llm_result.get("canonical_fields", {}))
            overall_conf = compute_overall_confidence(canonical)

            # Detect line item table and build with refs
            table_row = detect_table_start_row(meta["cell_map"])
            line_items = []
            if table_row:
                line_items = build_line_items_with_refs(file_bytes, name, table_row)

            results[name] = {
                "document_type": llm_result.get("document_type", "INVOICE"),
                "overall_confidence": overall_conf,
                "line_items_confidence": llm_result.get("line_items_confidence", 0.0),
                "canonical_fields": canonical,
                "line_items": remove_null_values(line_items),
            }
            processed += 1

        except Exception as e:
            logger.error(f"Error processing sheet '{name}': {e}")
            skipped.append({"name": name, "reason": str(e)})

    return {
        "_file_info": {
            "source_file": filename,
            "processed_at": datetime.now().isoformat(),
            "sheets_processed_as_invoices": processed,
            "sheets_skipped": len(skipped),
            "skipped_sheet_details": skipped
        },
        **results
    }


# ============================================================================
# 8. ASYNC ENTRY POINT (SSE streaming)
# ============================================================================
async def extract_invoice_from_bytes_with_status(file_bytes: bytes, filename: str):
    """Async generator — yields SSE status events then a final 'complete' event."""

    yield {"type": "status", "step": 1, "message": f"Reading Excel file: {filename}", "progress": 5}

    try:
        sheets = await asyncio.to_thread(extract_sheet_data, file_bytes, filename)
    except Exception as e:
        yield {"type": "error", "message": f"Failed to read file: {str(e)}"}
        return

    # Pre-filter
    sheets_to_process = []
    skipped = []

    yield {"type": "status", "step": 2, "message": "Scanning for invoice sheets...", "progress": 15}

    for name, meta in sheets.items():
        top_text = meta["cell_text"].lower()
        if any(kw in top_text for kw in ["invoice", "inv.", "commercial invoice", "commercial inv"]):
            sheets_to_process.append(name)
        else:
            skipped.append({"name": name, "reason": "No invoice keyword (pre-filter)"})

    if not sheets_to_process:
        yield {"type": "error", "message": "No invoice sheets found in this file."}
        return

    results = {}
    processed = 0
    total = len(sheets_to_process)

    for i, sheet_name in enumerate(sheets_to_process):
        progress = 20 + int((i / total) * 65)

        yield {"type": "status", "step": 3,
               "message": f"Extracting: '{sheet_name}' ({i+1}/{total})", "progress": progress}

        llm_result = await process_sheet_async(sheets[sheet_name])

        if not llm_result.get("is_invoice"):
            skipped.append({"name": sheet_name, "reason": f"LLM: {llm_result.get('document_type', 'Not Invoice')}"})
            yield {"type": "status", "step": 3,
                   "message": f"Sheet '{sheet_name}' - not an invoice", "progress": progress + 3}
            continue

        # Flatten canonical fields
        canonical = flatten_canonical_fields(llm_result.get("canonical_fields", {}))
        overall_conf = compute_overall_confidence(canonical)

        # Detect and build line items with refs (offloaded to thread)
        yield {"type": "status", "step": 3,
               "message": f"Building line items for '{sheet_name}'...", "progress": progress + 5}

        meta = sheets[sheet_name]
        table_row = detect_table_start_row(meta["cell_map"])
        line_items = []
        if table_row:
            line_items = await asyncio.to_thread(
                build_line_items_with_refs, file_bytes, sheet_name, table_row
            )

        results[sheet_name] = {
            "document_type": llm_result.get("document_type", "INVOICE"),
            "overall_confidence": overall_conf,
            "line_items_confidence": llm_result.get("line_items_confidence", 0.0),
            "canonical_fields": canonical,
            "line_items": remove_null_values(line_items),
        }
        processed += 1

        yield {"type": "status", "step": 3,
               "message": f"Sheet '{sheet_name}' - {len(line_items)} items | confidence: {overall_conf:.0%}",
               "progress": progress + 8}

    yield {"type": "status", "step": 4, "message": "Compiling results...", "progress": 95}

    final_output = {
        "_file_info": {
            "source_file": filename,
            "processed_at": datetime.now().isoformat(),
            "sheets_processed_as_invoices": processed,
            "sheets_skipped": len(skipped),
            "skipped_sheet_details": skipped
        },
        **results
    }

    yield {"type": "complete", "message": "Extraction successful", "data": final_output}